"""Ingestor de documentos — extrai texto, divide em chunks e cria skills.

Fluxo (todas as fontes):
  1. Extrair (title, raw_text) conforme source_type:
       - url   → reusa ``skill_importer.import_url`` (HTML→markdown)
       - text  → usa o conteúdo já recebido
       - pdf   → ``pypdf`` (lazy import)
       - xlsx  → ``openpyxl`` (lazy import)
  2. Calcular checksum sha256 do texto normalizado.
  3. Dividir em chunks de ~CHUNK_CHARS caracteres (com overlap simples).
  4. Para cada chunk: criar uma Skill (repository_id=<>, document_id=<>,
     chunk_index=i) e calcular embedding.
  5. Atualizar Document.status / chunks_count / indexed_at.

Reindexar = chamar ``reindex_document`` que apaga skills antigas (cascade do
FK), bumpa version e roda ingest novamente.
"""

from __future__ import annotations

import hashlib
import io
import logging
import re
import uuid
from typing import Any

from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.infrastructure.embeddings import embed_text
from app.infrastructure.orm_models import Document, Skill
from app.infrastructure.skill_importer import ImporterError, import_url

log = logging.getLogger(__name__)

CHUNK_CHARS = 2000
CHUNK_OVERLAP = 200


class IngesterError(RuntimeError):
    """Erro de ingestão (extração ou parsing)."""


# ── Extracção por tipo ─────────────────────────────────────────


async def _extract_url(uri: str) -> tuple[str, str]:
    try:
        out = await import_url(uri)
    except ImporterError as exc:
        raise IngesterError(f"falha a buscar URL: {exc}") from exc
    return out["title"], out["content"]


def _extract_text(content: str | None, fallback_title: str) -> tuple[str, str]:
    if not content or not content.strip():
        raise IngesterError("source_type='text' exige campo 'content' não vazio")
    title = fallback_title or content.strip().splitlines()[0][:120] or "Documento"
    return title, content


def _extract_pdf(blob: bytes, fallback_title: str) -> tuple[str, str]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise IngesterError(
            "pypdf não instalado — adicionar ao requirements.txt para suportar PDF"
        ) from exc
    try:
        reader = PdfReader(io.BytesIO(blob))
        pages = [page.extract_text() or "" for page in reader.pages]
    except Exception as exc:
        raise IngesterError(f"PDF inválido: {exc}") from exc
    text = "\n\n".join(p.strip() for p in pages if p.strip())
    title = (reader.metadata.title if reader.metadata else None) or fallback_title or "PDF"
    return str(title), text


def _extract_xlsx(blob: bytes, fallback_title: str) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngesterError(
            "openpyxl não instalado — adicionar ao requirements.txt para suportar XLSX"
        ) from exc
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as exc:
        raise IngesterError(f"XLSX inválido: {exc}") from exc

    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## {sheet_name}\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                parts.append(" | ".join(cells))
        parts.append("")
    text = "\n".join(parts).strip()
    return fallback_title or "Planilha", text


# ── Chunking simples ────────────────────────────────────────────


def chunk_text(text: str, size: int = CHUNK_CHARS, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Divide o texto em janelas de ``size`` chars com sobreposição leve.

    Tenta cortar em quebras de parágrafo/linha próximas para não partir frases.
    """
    text = re.sub(r"\n{3,}", "\n\n", text or "").strip()
    if not text:
        return []
    if len(text) <= size:
        return [text]

    chunks: list[str] = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + size, n)
        if end < n:
            window = text[start:end]
            cut = max(window.rfind("\n\n"), window.rfind("\n"), window.rfind(". "))
            if cut > size // 2:
                end = start + cut + 1
        chunks.append(text[start:end].strip())
        if end >= n:
            break
        start = max(end - overlap, start + 1)
    return [c for c in chunks if c]


def _checksum(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def _slugify(text: str, max_len: int = 80) -> str:
    text = (text or "").lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:max_len] or "doc"


# ── Pipeline principal ─────────────────────────────────────────


async def _delete_existing_chunks(session: AsyncSession, document_id: uuid.UUID) -> None:
    await session.execute(delete(Skill).where(Skill.document_id == document_id))


async def ingest_document(
    session: AsyncSession,
    document: Document,
    *,
    raw_blob: bytes | None = None,
    text_content: str | None = None,
) -> None:
    """Executa a ingestão de um documento (extracção + chunking + embeddings).

    O caller é responsável por commit/rollback. Em caso de erro, atualiza
    ``document.status='error'`` e ``error_message`` antes de relançar.
    """
    document.status = "processing"
    document.error_message = None
    await session.flush()

    try:
        title, raw_text = await _extract(document, raw_blob=raw_blob, text_content=text_content)
        if not raw_text.strip():
            raise IngesterError("documento sem texto extraível")

        if document.title in (None, "", "Documento"):
            document.title = title[:512]
        document.checksum = _checksum(raw_text)

        await _delete_existing_chunks(session, document.id)

        chunks = chunk_text(raw_text)
        base_slug = _slugify(document.title)
        for idx, chunk in enumerate(chunks):
            skill = Skill(
                id=uuid.uuid4(),
                repository_id=document.repository_id,
                document_id=document.id,
                chunk_index=idx,
                slug=await _ensure_unique_slug(session, f"{base_slug}-chunk-{idx}"),
                title=f"{document.title} (parte {idx + 1}/{len(chunks)})"[:512],
                summary=chunk[:300],
                content=chunk,
                tags=["support-doc", document.source_type],
                source_url=document.source_uri or None,
                active=True,
            )
            try:
                emb = await embed_text(f"{skill.title}\n\n{skill.content}"[:8000])
                if emb:
                    skill.embedding = emb
            except Exception as exc:
                log.warning("embedding falhou para chunk %d de %s: %s", idx, document.id, exc)
            session.add(skill)

        document.chunks_count = len(chunks)
        document.status = "indexed"
        document.indexed_at = _now()
    except Exception as exc:
        document.status = "error"
        document.error_message = str(exc)[:1000]
        log.exception("ingest falhou para document %s", document.id)
        raise


async def _extract(
    document: Document,
    *,
    raw_blob: bytes | None,
    text_content: str | None,
) -> tuple[str, str]:
    st = document.source_type
    fallback = document.title or ""
    if st == "url":
        return await _extract_url(document.source_uri)
    if st == "text":
        return _extract_text(text_content, fallback)
    if st == "pdf":
        if raw_blob is None:
            raise IngesterError("PDF sem blob — upload obrigatório")
        return _extract_pdf(raw_blob, fallback)
    if st == "xlsx":
        if raw_blob is None:
            raise IngesterError("XLSX sem blob — upload obrigatório")
        return _extract_xlsx(raw_blob, fallback)
    raise IngesterError(f"source_type não suportado: {st}")


async def _ensure_unique_slug(session: AsyncSession, base_slug: str) -> str:
    candidate = base_slug
    suffix = 2
    while True:
        existing = await session.scalar(select(Skill).where(Skill.slug == candidate))
        if not existing:
            return candidate
        candidate = f"{base_slug}-{suffix}"
        suffix += 1


def _now() -> Any:
    from datetime import UTC, datetime

    return datetime.now(UTC)
